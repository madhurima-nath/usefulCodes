# get necessary libraries
import pandas as pd
import openpyxl
import datetime
import numpy as np
import re


def check_url(string):
    """
    findall() used with the conditions
    which is valid for url in the string
    """
    regex = "http[s]?://(?:[a-zA-Z]|[0-9]|[$-_@.&+]|[!*\(\), ]|(?:%[0-9a-fA-F][0-9afA-F]))+"
    URL = re.findall(regex, string)
    if URL:
        return True
    else:
        return False


## parse excel file
def parse_excel(filename):
    """
    this function takes the excel file
    and parses it to extract the cell values
    where there is an update/change
    mentioned in the excel comments
    and/or hyperlinks in red.
    this will not work if
    - hyperlinks are not in red (RGB value is hard coded)
    - there are no excel comments added to the cells
    if there are updated links not in red,
    but the cell has a comment, the function
    will extrcat the information
    """

    # read excel file
    wb = openpyxl.load_workbook(filename)
    sheet = wb.active

    # create a dictionary to store all new comments
    comments_dict = {}

    # create a dictionary to store updated hyperlinks (in red)
    hyperlinks_dict = {}

    # get the maximum value for columns
    max_col = sheet.max_column

    for row in sheet.iter_rows(values_only=False):
        for cell in row:
            if cell.comment:
                comments_dict[f"{cell.coordinate}"] = [cell.coordinate,
                                                        sheet.cell(cell.row, 1).value,
                                                        (cell.comment.text.split("\n")[0].strip()),
                                                        cell.value,
                                                        sheet.cell(cell.row, cell.column-1).value,
                                                        sheet.cell(cell.row, cell.column-2).value,
                                                        sheet.cell(1, cell.column).value,
                                                        sheet.cell(cell.row, max_col).value]
                # ["cell_id", "country", "excel_comments", "cell_value",
                #    "prev_col1", "prev_col2", "excel_column_header", "comments"]

                # iterate through cells and extract only those hyperlinks in red
                if cell.font.color:
                    if "rgb" in cell.font.color.__dict__:
                        if cell.font.color.rgb == "FFFF0000": # rgb value for red font color
                            if cell.hyperlink:
                                # print(f"Cell {cell.coordinate} contains hyperlink")
                                hyperlinks_dict[f"{cell.coordinate}"] = cell.hyperlink.target

                # if updated hyperlinks are not in red, but cell has a comment
                # use the following tp get the hyperlinks_dict

                # check if the cell value is a hyperlink
                if ( (type(cell.value) == str)
                    and (check_url(cell.value) == True) ):
                    hyperlinks_dict[f"{cell.coordinate}"] = cell.value

    df = pd.DataFrame(comments_dict.items(), columns=["cell_id", "cell_list"])
    return df, hyperlinks_dict


def add_update_text_col(df):
    """
    this section parses different columns
    to get only the updates from the excel cells
    where the updates are mentioned in numbers.

    due to multiple variations of the updates,
    different if statements are used to capture
    the possibilities.
    """

    df["update_text"] = ""

    for i in range(len(df)):
        if df.loc[i, "updated_link"] == False:
            # get latest updates from excel cell value into update_text column
            if ( (type(df.loc[i, "cell_value"]) == str)
                and (type(df.loc[i, "update_num"]) == str) ):
                split_num = df.loc[i, "update_num"]
                if split_num in df.loc[i, "cell_value"]:
                    split_list = df.loc[i, "cell_value"].split(split_num)

                    # sometimes the updates are made to the start
                    # of the cell entry and at other times it could
                    # be in between the list
                    for x in split_list:
                        if split_num not in x and "\n" not in x:
                            df.loc[i, "update_text"] = x.strip()
                        elif split_num not in x and "\n" in x:
                            df.loc[i, "update_text"] = x.split("\n")[0].strip()

            # get the dates from excel cell values into update_text column if in datetime format
            elif (type(df.loc[i, "cell_value"]) == datetime.datetime
                and "date" in df.loc[i, "excel_comments"]):
                df.loc[i, "update_text"] = df.loc[i, "cell_value"]

            # get all updated dates from excel cell values into update_text column if there are multiple dates
            elif (type(df.loc[i, "cell_value"]) == str
                and ("dates" or "both" or "all") in df.loc[i, "excel_comments"]):
                df.loc[i, "update_text"] = df.loc[i, "cell_value"].strip()

    return df


def extract_previous_col_text(df):
    """
    this section parses the dateframe
    to get the information from the previous
    columns. 
    due to multiple variations of updates,
    multiple if statements are used.
    """

    df["adj_prev_col1"] = ""
    df["adj_prev_col2"] = ""

    for i in range(len(df)):
        if ( ("date" in df.loc[i, "excel_comments"])
            and (type(df.loc[i, "cell_value"]) == str)
            and (("dates" or "both" or "all") not in df.loc[i, "excel_comments"]) ):
            if ( (df.loc[i, "update_num"] != np.nan)
                and (type(df.loc[i, "update_num"]) == str) ):
                split_num = df.loc[i, "update_num"]
                if split_num in df.loc[i, "cell_value"]:
                    split_list1 = df.loc[i, "prev_col1"].split(split_num)
                    split_list2 = df.loc[i, "prev_col2"].split(split_num)
                    # sometimes the updates are made to the start
                    # of the cell value and at other times it is in between
                    for x in split_list1:
                        if split_num not in x and "\n" not in x:
                            df.loc[i, "adj_prev_col1"] = x.strip()
                        elif split_num not in x and "\n" in x:
                            df.loc[i, "adj_prev_col1"] = x.split("\n")[0].strip()
                    
                    for x in split_list2:
                        if split_num not in x and "\n" not in x:
                            df.loc[i, "adj_prev_col2"] = x.strip()
                        elif split_num not in x and "\n" in x:
                            df.loc[i, "adj_prev_col2"] = x.split("\n")[0].strip()

            else:
                df.loc[i, "adj_prev_col1"] = df.loc[i, "prev_col1"]
                df.loc[i, "adj_prev_col2"] = df.loc[i, "prev_col2"]
        
        elif ( ("date" in df.loc[i, "excel_comments"])
            and (type(df.loc[i, "cell_value"]) == str)
            and (("dates" or "both" or "all") in df.loc[i, "excel_comments"]) ):
            df.loc[i, "adj_prev_col1"] = df.loc[i, "prev_col1"]
            df.loc[i, "adj_prev_col2"] = df.loc[i, "prev_col2"]

        elif ( ("date" in df.loc[i, "excel_comments"])
            and (type(df.loc[i, "cell_value"]) == datetime.datetime) ):
            df.loc[i, "adj_prev_col1"] = df.loc[i, "prev_col1"]
            df.loc[i, "adj_prev_col2"] = df.loc[i, "prev_col2"]

    return df


def process_df(filename):
    """
    processes initial dataframe obtained
    from reading the excel file.

    add new columns after parsing the contents
    of excel cells.
    """

    # parse excel to extract updates
    df1, hyperlinks_dict = parse_excel(filename)
    df2 = pd.DataFrame(df1["cell_list"].to_list(),
            columns=["cell_id", "country", "excel_comments",
                    "cell_value", "prev_col1", "prev_col2", 
                    "excel_column_header", "comments"], 
                    index=df1.index)

    # check for updated hyperlinks
    df2["updated_link"] = df2["cell_value"].isin(hyperlinks_dict.values())
    # extract only the numbers (reference numbers) from excel_comments column
    df2["update_num"] = df2["excel_comments"].str.findall(r"\d+")

    df2["column_header"] = ""

    # get column headers for non-null update_text column
    # if the column header is null/missing, then write Updated Value (Confirm)
    for i in range(len(df2)):
        if (df2.loc[i, "excel_column_header"] != None):
            df2.loc[i, "column_header"] = df2.loc[i, "excel_column_header"]
        elif ( (df2.loc[i, "excel_column_header"] == None)
            # and ("date" not in df2.loc[i, "excel_column_header"])
            and (df2.loc[i, "updated_link"] == False) ):
            df2.loc[i, "column_header"] = "Updated Value (Confirm)"
        else:
            # for weblinks, dates etc (where Effective Date is not column header)
            df2.loc[i, "column_header"] = "Does not apply"

    multi_dict = {}

    for i in range(len(df2)):
        num_list = []
        temp = df2.loc[i, "update_num"]

        # this section adds a `.` at the end of the
        # numeric values to be later used to parse
        # the string/text to only get the updated ones
        for x in temp:
            # making sure numbers from hyperlinks do not
            # impact the output
            if "no change" not in df2.loc[i, "excel_comments"].lower():
                num_list.append(x+".")
            else:
                pass

        multi_dict[i] = {"country": df2.loc[i, "country"],
                        "excel_comments": df2.loc[i, "excel_comments"],
                        "cell_value": df2.loc[i, "cell_value"],
                        "comments": df2.loc[i, "comments"],
                        "updated_link": df2.loc[i, "updated_link"],
                        "column_header": df2.loc[i, "column_header"],
                        "prev_col1": df2.loc[i, "prev_col1"],
                        "prev_col2": df2.loc[i, "prev_col2"],
                        "update_num": num_list}
        
    df3 = pd.DataFrame.from_dict(multi_dict, orient="index")

    # replace the rows in original dataframe containing multiple update_num
    df2.update(df3)
    df4 = df2.explode("update_num")
    df4 = df4.reset_index(drop=True)

    df5 = add_update_text_col(df4)
    df5 = extract_previous_col_text(df5)

    return df5


## process dataframe
def get_processed_df(filename):
    """
    this function returns the processed dataframe
    """
    df = process_df(filename=filename)
    return df